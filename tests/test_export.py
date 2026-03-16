from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import cast

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from airq_mcp_timeseries.errors import UnsupportedOutputFormatError
from airq_mcp_timeseries.models import (
    ExportResult,
    HistoryQuery,
    MetricInfo,
    Selector,
    SeriesPoint,
    SeriesSet,
    TimeSeries,
)
from airq_mcp_timeseries.services.export import export_history, export_series_set


def test_export_series_set_writes_csv(sample_metrics, sample_series_set) -> None:
    result = export_series_set(sample_series_set, output_format="csv", metric_info=sample_metrics[0])

    assert result == ExportResult(
        output_format="csv",
        mime_type="text/csv; charset=utf-8",
        payload=result.payload,
    )
    assert isinstance(result.payload, str)
    lines = result.payload.strip().splitlines()
    assert lines[0] == "timestamp,series,metric,unit,value"
    assert "Wohnzimmer,pm2_5,ug/m3,0.0" in lines[1]


def test_export_series_set_writes_xlsx(sample_metrics, sample_series_set) -> None:
    result = export_series_set(sample_series_set, output_format="xlsx", metric_info=sample_metrics[0])

    assert result.output_format == "xlsx"
    assert result.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert isinstance(result.payload, bytes)

    workbook = load_workbook(filename=BytesIO(result.payload))
    worksheet = cast(Worksheet, workbook.active)
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    assert rows[0] == ("timestamp", "series", "metric", "unit", "value")
    assert rows[1][1:] == (
        "Wohnzimmer",
        "pm2_5",
        "ug/m3",
        0,
    )


def test_export_series_set_falls_back_to_metric_unit_when_series_unit_missing() -> None:
    series_set = SeriesSet(
        metric="co2",
        series=[
            TimeSeries(
                id="wohnzimmer",
                label="Wohnzimmer",
                unit=None,
                points=[SeriesPoint(ts="2026-03-01T00:00:00+00:00", value=650.0)],
            )
        ],
        start="2026-03-01T00:00:00+00:00",
        end="2026-03-01T00:00:00+00:00",
        source_resolution_s=300,
    )

    result = export_series_set(
        series_set,
        output_format="csv",
        metric_info=MetricInfo(key="co2", label="CO2", unit="ppm"),
    )

    assert isinstance(result.payload, str)
    assert "Wohnzimmer,co2,ppm,650.0" in result.payload


def test_export_series_set_keeps_multiple_series_in_one_csv() -> None:
    series_set = SeriesSet(
        metric="co2",
        series=[
            TimeSeries(
                id="wohnzimmer",
                label="Wohnzimmer",
                unit="ppm",
                points=[SeriesPoint(ts="2026-03-01T00:00:00+00:00", value=650.0)],
            ),
            TimeSeries(
                id="buero",
                label="Büro",
                unit="ppm",
                points=[SeriesPoint(ts="2026-03-01T00:00:00+00:00", value=700.0)],
            ),
        ],
        start="2026-03-01T00:00:00+00:00",
        end="2026-03-01T00:00:00+00:00",
        source_resolution_s=300,
    )

    result = export_series_set(series_set, output_format="csv")

    assert isinstance(result.payload, str)
    lines = result.payload.strip().splitlines()
    assert len(lines) == 3
    assert "Wohnzimmer,co2,ppm,650.0" in lines[1]
    assert "Büro,co2,ppm,700.0" in lines[2]


@pytest.mark.asyncio
async def test_export_history_resamples_when_requested(sample_metrics, sample_series_set) -> None:
    class Provider:
        async def get_capabilities(self):
            from airq_mcp_timeseries.models import CapabilitySet

            return CapabilitySet(latest_values=True, historical_values=True)

        async def list_metrics(self, selector=None):
            return sample_metrics

        async def get_history(self, query):
            return sample_series_set

    query = HistoryQuery(
        selector=Selector(devices=["Wohnzimmer"]),
        metric="PM2.5",
        start=datetime(2026, 3, 1, tzinfo=UTC),
        end=datetime(2026, 3, 1, 1, 0, 0, tzinfo=UTC),
        aggregation="mean",
        requested_interval_s=1800,
    )

    result = await export_history(Provider(), query, output_format="csv")

    assert isinstance(result.payload, str)
    lines = result.payload.strip().splitlines()
    assert len(lines) == 3
    assert lines[1].endswith(",2.5")
    assert lines[2].endswith(",8.5")


def test_export_series_set_rejects_unknown_format(sample_series_set) -> None:
    with pytest.raises(UnsupportedOutputFormatError):
        export_series_set(sample_series_set, output_format="json")
